from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
count=0
average_salary =0
#write your code here
for rinda in range (2,ws.max_row+1) :
    stundas = ws["C"+ str(rinda)].value
    likme=ws["B"+ str(rinda)].value
    if isinstance(stundas,(int,float)) and isinstance(likme,(int,float)):
        alga=stundas*likme
        if alga>3000:
            total+=1
            count+=alga
#print(count)
#print(total)
average_salary = count/total
average_salary2 = round(average_salary, 0)
print('Videja alga:',average_salary2)
wb.close()
